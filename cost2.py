# -*- coding: UTF-8 -*-
from openpyxl import load_workbook

from openpyxl.worksheet.filters import CustomFilterValueDescriptor


def monkey_set(self, instance, value):
    pass


CustomFilterValueDescriptor.__set__ = monkey_set


def process3():
    targetWb = load_workbook("./DBU-1-6-cost.xlsx")
    referWb = load_workbook("./cost-refer-1-6.xlsx")
    targetColumn = 14

    sheets = ["预算汇总表-高科技", "预算汇总表-金融", "预算汇总表-大客户", "预算汇总表-业务拓展",
              "预算汇总表-产品解决方案", "预算汇总表-生产部", "预算汇总表-技术支持", "预算汇总表-探针", "预算汇总表-至安盾", "预算汇总表-至安盾二部", "预算汇总表-业务支持"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("cost_output1-6.xlsx")


def process2():
    targetWb = load_workbook("./DBU-2023-total.xlsx")
    referWb = load_workbook("./2023-q1.xlsx")
    targetColumn = 7

    sheets = ["DBU汇总", "高科技销售部", "金融销售部", "政法销售部",
              "军工和大客户销售部", "通信销售部", "业务拓展组", "产品部", "解决方案一部", "解决方案三部", "生产部", "技术支持部", "产品开发一部", "产品开发二部", "产品开发三部", "技术研究部", "安全产品事业部-业务支持"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("DBU_2023_cost_output.xlsx")


def processDBU2022():
    targetWb = load_workbook("./dbu-12.xlsx")
    referWb = load_workbook("./2022-dbu.xlsx")
    targetColumn = 39

    sheets = ["预算汇总表-高科技", "预算汇总表-金融", "预算汇总表-大客户", "预算汇总表-业务拓展",
              "预算汇总表-产品解决方案", "预算汇总表-技术支持", "预算汇总表-探针", "预算汇总表-至安盾1", "预算汇总表-至安盾2", "预算汇总表-业务支持", "预算汇总表-生产部"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("cost_output_dbu_2022.xlsx")


def process1():
    targetWb = load_workbook("./2023-total.xlsx")
    referWb = load_workbook("./2023-q1.xlsx")
    targetColumn = 7

    sheets = ["PBU", "DBU", "MBU", "研究院", "研发管理部", "项目管理与实施部",
              "总裁办", "市场部", "战略研究部", "财务+内审", "人事行政中心", "董办"]

    generate(sheets, targetColumn, targetWb, referWb)
    # write out
    targetWb.save("2023_cost_output.xlsx")


def generate(sheets, targetColumn, targetWb, referWb):
    for sheet in sheets:
        print("process sheet:", sheet)
        processDepart(targetWb[sheet], targetColumn, referWb)


def processDepart(sheet, targetColumn, referWb):
    typeColumn = 2
    departPool = set(sheet.cell(row=1, column=2).value.split("、"))
    for i in range(5, sheet.max_row + 1):
        typeCell = sheet.cell(row=i, column=typeColumn)
        if isEmptyCell(typeCell):
            continue
        # print("typeCell vaule:", typeCell.value)
        # make sure typeCell.value to string

        types = typeConvert(str(typeCell.value).split("、"))
        if typeCell.value == "500104…550109、660104…660109、660204…660209、660404…660409、28010204…28010209":
            print("types:", types)

        converts = []
        for item in types:
            if isinstance(item, list):
                converts.extend(item)
            else:
                converts.append(item)
        # print("converted:", converts)
        count = computeTypesCost(converts, referWb, departPool)
        # print("process row:", i, count)
        sheet.cell(row=i, column=targetColumn).value = round(
            count, 2)


def typeConvert(types):
    return list(map(checkRange, types))


def checkRange(item):
    items = item.split("…")
    if len(items) > 1:
        list = []
        for i in range(0, int(items[1]) - int(items[0]) + 1):
            list.append(str(int(items[0]) + i))
        return list
    else:
        return item


def computeTypesCost(types, referWb, departs):
    config = {
        "生产成本": {
            "typeIndex": 2,
            "departIndex": 4,
            "amountIndex": 6,
        },
        "销售费用": {
            "typeIndex": 2,
            "departIndex": 4,
            "amountIndex": 6,
        },
        "管理费用": {
            "typeIndex": 2,
            "departIndex": 4,
            "amountIndex": 6,
        },
        "研发费用": {
            "typeIndex": 2,
            "departIndex": 4,
            "amountIndex": 6,
        }
    }

    typeSet = set(types)
    count = 0.0
    # gothrough refer sheer
    for i in range(0, len(referWb.sheetnames)):
        sheetName = referWb.sheetnames[i]
        if sheetName in config:
            c = config[sheetName]
            # print("process sheet:", sheetName)
            count += processSheet(referWb[sheetName], c["typeIndex"], c["departIndex"],
                                  c["amountIndex"], typeSet, departs, sheetName)

    return count


def processSheet(sheet, typeIndex, departIndex, amountIndex, types, departs, sheetName):
    count = 0.0
    for i in range(2, sheet.max_row + 1):
        departCell = sheet.cell(row=i, column=departIndex)
        if isEmptyCell(departCell):
            continue
        typeCell = sheet.cell(row=i, column=typeIndex)
        if isEmptyCell(typeCell):
            continue
        amountCell = sheet.cell(row=i, column=amountIndex)
        if isEmptyCell(amountCell):
            continue

        typeValue = str(typeCell.value).replace("'", "")
        departValue = str(departCell.value).replace("'", "")
        if departValue in departs and typeValue in types:
            print("found match:", sheetName, departValue,
                  typeValue, amountCell.value)
            count += float(amountCell.value)

    # print("sheet process count:", count)
    return count


def isEmptyCell(cell):
    return cell.value is None or cell.value == ""


process1()
