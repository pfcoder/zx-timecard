# coding=utf-8
from openpyxl import load_workbook
import sys

projects_info = {}

# this set of users use 7 as 8
specials = {"顾明璇", "田玲杰", "张孟祎", "郝跃红"}
# excludes = {"李庆乐"}
excludes = {}
ignore_salary = {}


def init():
    initial = {
        "应发工资": 0.0,
        "养老保险_个人": 0.0,
        "医疗保险_个人": 0.0,
        "失业保险_个人": 0.0,
        "大病医疗保险_个人": 0.0,
        "住房公积金_个人": 0.0,
        "养老保险_公司": 0.0,
        "医疗保险_公司": 0.0,
        "失业保险_公司": 0.0,
        "工伤保险_公司": 0.0,
        "生育保险_公司": 0.0,
        "大病医疗保险_公司": 0.0,
        "住房公积金_公司": 0.0,
        "所得税": 0.0,
        "实发工资": 0.0
    }

    sheet = load_workbook("./input.xlsx")

    # build salary map info
    salary_records = setupSalary(sheet)
    # build attend map info
    attend_records = setupAttend(sheet)

    # go through process working time
    time_info, projects = processAllTime(sheet, salary_records)
    result, prjDepart, departTimes, departCost = processSalary(
        time_info, salary_records, attend_records, initial, projects)

    update(sheet, result, initial, prjDepart, departTimes, departCost)
    summaryTime(projects, time_info, sheet, departTimes, departCost)
    verify(time_info, salary_records, result, attend_records, departCost)
    print("process done!")


def isEmptyCell(cell):
    return cell.value is None or cell.value == ""


# setup a employee name to depart and salary infos mapping


def setupSalary(wb):
    salary = {}
    salarySheet = wb["工资"]
    # print("Salary info length:", salarySheet.max_row)
    for i in range(2, salarySheet.max_row + 1):
        nameCell = salarySheet.cell(row=i, column=1)
        if isEmptyCell(nameCell):
            print("salary empty name cell:", i)
            continue
        print("process:", nameCell.value)
        departCell = salarySheet.cell(row=i, column=8)
        if isEmptyCell(departCell):
            print("salary empty depart cell:", i)
            continue
        # print("process salary:", nameCell.value, i,
        #      salarySheet.cell(row=1, column=3).value)

        # check if name already exist
        if nameCell.value in salary:
            print("salary name exist:", nameCell.value)
            # panic
            sys.exit(1)

        salary[nameCell.value] = {
            "depart": departCell.value,
            "应发工资": float(salarySheet.cell(row=i, column=11).value),
            "养老保险_个人": float(salarySheet.cell(row=i, column=12).value),
            "医疗保险_个人": float(salarySheet.cell(row=i, column=13).value),
            "失业保险_个人": float(salarySheet.cell(row=i, column=14).value),
            "大病医疗保险_个人": float(salarySheet.cell(row=i, column=15).value),
            "住房公积金_个人": float(salarySheet.cell(row=i, column=16).value),
            "养老保险_公司": float(salarySheet.cell(row=i, column=18).value),
            "医疗保险_公司": float(salarySheet.cell(row=i, column=19).value),
            "失业保险_公司": float(salarySheet.cell(row=i, column=20).value),
            "工伤保险_公司": float(salarySheet.cell(row=i, column=21).value),
            "生育保险_公司": float(salarySheet.cell(row=i, column=22).value),
            "大病医疗保险_公司": float(salarySheet.cell(row=i, column=23).value),
            "住房公积金_公司": float(salarySheet.cell(row=i, column=24).value),
            "所得税": float(salarySheet.cell(row=i, column=26).value),
            "实发工资": float(salarySheet.cell(row=i, column=27).value),
        }

    # print(salary)
    return salary


# compute all employee's attendance hours
# sheet column 11 is attendance days of current month
# return a map of employee name to attendance hours(month)


def setupAttend(wb):
    attendSheet = wb["考勤"]
    result = {}

    for i in range(2, attendSheet.max_row + 1):
        nameCell = attendSheet.cell(row=i, column=1)
        name = nameCell.value
        print("attend name:", name)
        result[name] = float(attendSheet.cell(row=i, column=11).value) * 8

    return result


def processAllTime(wb, salary):
    timeRecordsSheet = wb["工时"]
    # name as key
    result = {}
    # project id as key
    projects = {}

    missSalaryNames = {}

    # print(timeRecordsSheet.cell(row=2, column=2).value)
    # go through timeRecordsSheet row by row
    for i in range(2, timeRecordsSheet.max_row + 1):
        nameCell = timeRecordsSheet.cell(row=i, column=2)
        # print("process:", nameCell.value)
        if isEmptyCell(nameCell):
            print("empty name cell:", i)
            continue

        name = nameCell.value
        # print(name)
        if name in excludes:
            print("exclude:", name)
            continue

        if name not in salary:
            print("name not in salary", name)
            missSalaryNames[name] = 1
            continue

        depart = salary[name]["depart"]

        projectIdCell = timeRecordsSheet.cell(row=i, column=8)
        if isEmptyCell(projectIdCell):
            print("empty project cell:", i, name)
            continue

        if name not in result:
            result[name] = {
                "total_hours": 0,
                "projects": {}
            }

        hoursCell = timeRecordsSheet.cell(row=i, column=12)
        if isEmptyCell(hoursCell):
            print("empty hours cell:", i)
            continue
        hours = int(hoursCell.value)
        if name in specials and hours == 7:
            print("special hours change to 8:", name, hours)
            hours = 8

        result[name]["total_hours"] += hours

        projectNameCell = timeRecordsSheet.cell(row=i, column=9)
        if projectIdCell.value not in projects_info:
            projects_info[projectIdCell.value] = projectNameCell.value

        if projectIdCell.value not in result[name]["projects"]:
            result[name]["projects"][projectIdCell.value] = 0

        result[name]["projects"][projectIdCell.value] += hours

        if projectIdCell.value not in projects:
            # employee name as key (should use employee id, TODO)
            projects[projectIdCell.value] = {}

        if name not in projects[projectIdCell.value]:
            projects[projectIdCell.value][name] = 0
        projects[projectIdCell.value][name] += hours

    print("miss salary names:", missSalaryNames.keys())
    return (result, projects)


def processSalary(timeInfo, salaryInfo, attendRecord, initial, projects):
    # for name in timeInfo:
    # project id as key
    result = {}
    # standardHours = workingDays * 8
    prjDepart = {}
    departTimes = {}
    departCostRecords = {}
    # go through time info
    for name in timeInfo:
        if name not in salaryInfo:
            print("can not find salary:", name)
            continue
        salary = salaryInfo[name]
        totalHours = timeInfo[name]["total_hours"]

        if totalHours == 0:
            print("name 0 hours:", name)
            ignore_salary[name] = 1
            continue

        departHours = 0
        depart = salary["depart"]

        if name not in attendRecord:
            print("can not find attend:", name)
            continue

        standardHours = attendRecord[name]

        departCost = initial.copy()
        if totalHours < standardHours:
            departHours = standardHours - totalHours

            if depart not in departCostRecords:
                departCostRecords[depart] = initial.copy()

            for key in departCostRecords[depart]:
                departCost[key] = salary[key] * (departHours / standardHours)
                departCostRecords[depart][key] += departCost[key]

            if depart not in departTimes:
                departTimes[depart] = {}
            if name not in departTimes[depart]:
                departTimes[depart][name] = 0
            departTimes[depart][name] += departHours

        for project in timeInfo[name]["projects"]:
            if project not in result:
                result[project] = initial.copy()
            for key in result[project]:
                prjCost = (salary[key] - departCost[key]) * \
                    (timeInfo[name]["projects"][project] / totalHours)
                result[project][key] += prjCost
                if project not in prjDepart:
                    prjDepart[project] = {}
                if depart not in prjDepart[project]:
                    prjDepart[project][depart] = {}
                if key not in prjDepart[project][depart]:
                    prjDepart[project][depart][key] = 0.0
                prjDepart[project][depart][key] += prjCost
    # print(departCostRecords)
    return (result, prjDepart, departTimes, departCostRecords)
    # print(result)


def update(wb, records, initial, prjDepart, departTimes, departCostRecords):
    # write to excel
    sheetName = "time_cost"
    if sheetName in wb.sheetnames:
        del wb[sheetName]
    targetSheet = wb.create_sheet(sheetName)
    titles = ["项目代号", "项目名称", "部门"] + list(initial.keys())
    # print(titles)
    targetSheet.append(titles)

    for project in records:
        prjCode = project
        prjName = None
        if project in projects_info:
            # rowShare.append(projects_info[project])
            prjName = projects_info[project]
        # else:
        #    row.append(None)
        # row.append(None)
        # for i in range(3, len(titles)):
        #    row.append(round(records[project][titles[i]], 2))
        # targetSheet.append(row)
        # append depart detail
        if project in prjDepart:
            departCost = prjDepart[project]
            for depart in departCost:
                detailRow = [prjCode, prjName, depart]
                for j in range(3, len(titles)):
                    detailRow.append(round(departCost[depart][titles[j]], 2))
                targetSheet.append(detailRow)

    for prj in departCostRecords:
        print("project:", prj)
        row = [prj]
        if prj in projects_info:
            row.append(projects_info[prj])
        else:
            row.append(None)
        row.append(None)
        for i in range(3, len(titles)):
            row.append(round(departCostRecords[prj][titles[i]], 2))
        targetSheet.append(row)

    wb.save("output.xlsx")


def summaryTime(prjInfo, timeInfo, wb, departTimes, departCostRecords):
    sheetName = "time_summary"
    if sheetName in wb.sheetnames:
        del wb[sheetName]
    targetSheet = wb.create_sheet(sheetName)
    # x is employee names, y is project nos
    # build first title row
    titles = ['项目'] + list(timeInfo.keys()) + ['汇总']
    targetSheet.append(titles)
    total = [0] * len(titles)
    total[0] = '工时汇总'
    for project in prjInfo:
        row = [project]
        prj = prjInfo[project]
        prjTotal = 0
        for i in range(1, len(titles) - 1):
            name = titles[i]
            if name in prj:
                row += [prj[name]]
                total[i] += prj[name]
                prjTotal += prj[name]
            else:
                row += [0]
        row += [prjTotal]
        targetSheet.append(row)

    for project in departTimes:
        row = [project]
        prj = departTimes[project]
        prjTotal = 0
        for i in range(1, len(titles) - 1):
            name = titles[i]
            if name in prj:
                row += [prj[name]]
                total[i] += prj[name]
                prjTotal += prj[name]
            else:
                row += [0]
        row += [prjTotal]
        targetSheet.append(row)

    targetSheet.append(total)

    wb.save("output.xlsx")


def verify(timeInfo, salary, processResult, attendRecord, departCost):
    print("start verify....")
    sTotal = 0.0
    eCount = 0
    for name in timeInfo:
        if name in salary and name not in ignore_salary:
            sTotal += salary[name]["应发工资"]
            eCount += 1
    print("included employee salary total:", sTotal, eCount)

    pTotal = 0
    for prj in processResult:
        pTotal += processResult[prj]["应发工资"]

    pDepartTotal = 0
    for prj in departCost:
        pDepartTotal += departCost[prj]["应发工资"]

    print("total project spend:", pTotal + pDepartTotal, pTotal, pDepartTotal)

    # cross verify project cost compute

    verifyResult = {}
    # standardHours = workDays * 8
    for project in processResult:
        pTotal = 0.0
        for name in timeInfo:
            if name in ignore_salary:
                continue
            if project in timeInfo[name]["projects"]:
                if name not in verifyResult:
                    verifyResult[name] = {
                        "salary": salary[name]["应发工资"],
                        "total_hours": timeInfo[name]["total_hours"],
                        "project_cost": 0.0,
                        "depart_cost": 0.0,
                        "project_detail": {}
                    }
                prjHours = timeInfo[name]["projects"][project]
                totalHours = timeInfo[name]["total_hours"]
                standardHours = attendRecord[name]
                if totalHours < standardHours:
                    totalHours = standardHours
                prjRatio = prjHours / totalHours
                prjCostShare = salary[name]["应发工资"] * prjRatio
                pTotal += prjCostShare
                verifyResult[name]["project_cost"] += prjCostShare
                verifyResult[name]["project_detail"][project] = {
                    "hours": prjHours,
                    "cost": prjCostShare
                }

        processResult[project]["verified_cost"] = pTotal
        # print("verify:", project, processResult[project]["应发工资"], pTotal)


init()
