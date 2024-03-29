# -*- coding: UTF-8 -*-
from openpyxl import load_workbook

from openpyxl.worksheet.filters import CustomFilterValueDescriptor


def monkey_set(self, instance, value):
    pass


CustomFilterValueDescriptor.__set__ = monkey_set


def process3():
    targetWb = load_workbook("./DBU-1-6-cost.xlsx")
    referWb = load_workbook("./cost-refer-1-6.xlsx")
    targetColumn = 14

    sheets = ["预算汇总表-高科技", "预算汇总表-金融", "预算汇总表-大客户", "预算汇总表-业务拓展",
              "预算汇总表-产品解决方案", "预算汇总表-生产部", "预算汇总表-技术支持", "预算汇总表-探针", "预算汇总表-至安盾", "预算汇总表-至安盾二部", "预算汇总表-业务支持"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("cost_output1-6.xlsx")


def process2():
    targetWb = load_workbook("./DBU-cost.xlsx")
    referWb = load_workbook("./cost_dbu_refer_12.xlsx")
    targetColumn = 35

    sheets = ["预算汇总表-高科技", "预算汇总表-金融", "预算汇总表-大客户", "预算汇总表-业务拓展",
              "预算汇总表-产品解决方案", "预算汇总表-技术支持", "预算汇总表-探针", "预算汇总表-至安盾1", "预算汇总表-至安盾2", "预算汇总表-业务支持", "预算汇总表-生产部"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("cost_output2.xlsx")


def processDBU2022():
    targetWb = load_workbook("./dbu-12.xlsx")
    referWb = load_workbook("./2022-dbu.xlsx")
    targetColumn = 39

    sheets = ["预算汇总表-高科技", "预算汇总表-金融", "预算汇总表-大客户", "预算汇总表-业务拓展",
              "预算汇总表-产品解决方案", "预算汇总表-技术支持", "预算汇总表-探针", "预算汇总表-至安盾1", "预算汇总表-至安盾2", "预算汇总表-业务支持", "预算汇总表-生产部"]
    generate(sheets, targetColumn, targetWb, referWb)

    # write out
    targetWb.save("cost_output_dbu_2022.xlsx")


def process1():
    targetWb = load_workbook("./cost_total.xlsx")
    referWb = load_workbook("./2022.xlsx")
    targetColumn = 38

    sheets = ["汇总-DBU", "预算汇总表MBU", "汇总表-PBU", "预算汇总表-业务拓展部", "预算汇总表-商务拓展部", "预算汇总表-研究院",
              "预算汇总表-总裁办合并", "预算汇总表-市场部", "预算汇总表-人事行政中心", "预算汇总表-董办", "预算汇总表-财务中心+法务"]

    generate(sheets, targetColumn, targetWb, referWb)
    # write out
    targetWb.save("cost_output.xlsx")


def generate(sheets, targetColumn, targetWb, referWb):
    for sheet in sheets:
        processDepart(targetWb[sheet], targetColumn, referWb)


def processDepart(sheet, targetColumn, referWb):
    typeColumn = 4
    departPool = set(sheet.cell(row=1, column=1).value.split("、"))
    for i in range(5, sheet.max_row + 1):
        typeCell = sheet.cell(row=i, column=typeColumn)
        if isEmptyCell(typeCell):
            continue
        types = typeConvert(typeCell.value.split("、"))
        converts = []
        for item in types:
            if isinstance(item, list):
                converts.extend(item)
            else:
                converts.append(item)
        # print("converted:", converts)
        count = computeTypesCost(converts, referWb, departPool)
        print("process row:", i, count)
        sheet.cell(row=i, column=targetColumn).value = round(
            count, 2)


def typeConvert(types):
    return list(map(checkRange, types))


def checkRange(item):
    items = item.split("…")
    if len(items) > 1:
        list = []
        for i in range(0, int(items[1]) - int(items[0]) + 1):
            list.append(str(int(items[0]) + i))
        return list
    else:
        return item


def computeTypesCost(types, referWb, departs):
    config = {
        "收入": {
            "typeIndex": 2,
            "departIndex": 7,
            "amountIndex": 10,
        },
        "生产成本": {
            "typeIndex": 1,
            "departIndex": 4,
            "amountIndex": 5,
        },
        "销售费用": {
            "typeIndex": 2,
            "departIndex": 5,
            "amountIndex": 6,
        },
        "管理费用": {
            "typeIndex": 2,
            "departIndex": 5,
            "amountIndex": 6,
        },
        "研发费用": {
            "typeIndex": 2,
            "departIndex": 5,
            "amountIndex": 6,
        }
    }

    typeSet = set(types)
    count = 0.0
    # gothrough refer sheer
    for i in range(0, len(referWb.sheetnames)):
        sheetName = referWb.sheetnames[i]
        if sheetName in config:
            c = config[sheetName]
            # print("process sheet:", sheetName)
            count += processSheet(referWb[sheetName], c["typeIndex"], c["departIndex"],
                                  c["amountIndex"], typeSet, departs, sheetName)

    return count


def processSheet(sheet, typeIndex, departIndex, amountIndex, types, departs, sheetName):
    count = 0.0
    for i in range(2, sheet.max_row + 1):
        departCell = sheet.cell(row=i, column=departIndex)
        if isEmptyCell(departCell):
            continue
        typeCell = sheet.cell(row=i, column=typeIndex)
        if isEmptyCell(typeCell):
            continue
        amountCell = sheet.cell(row=i, column=amountIndex)
        if isEmptyCell(amountCell):
            continue

        typeValue = str(typeCell.value).replace("'", "")
        departValue = str(departCell.value).replace("'", "")
        if departValue in departs and typeValue in types:
            print("found match:", sheetName, departValue,
                  typeValue, amountCell.value)
            count += float(amountCell.value)

    # print("sheet process count:", count)
    return count


def isEmptyCell(cell):
    return cell.value is None or cell.value == ""


processDBU2022()
