# coding=utf-8
from openpyxl import load_workbook
import sys

from openpyxl.worksheet.filters import CustomFilterValueDescriptor


def monkey_set(self, instance, value):
    pass


CustomFilterValueDescriptor.__set__ = monkey_set

projects_info = {}
initial = {}

# this set of users use 7 as 8
# specials = {"顾明璇", "田玲杰", "张孟祎", "郝跃红"}
specials = {}
# excludes = {"李庆乐"}
excludes = {}


def init():
    print("start")
    sheet = load_workbook("./year.xlsx")
    print("load done")
    # build salary map info
    salary_records = setupSalary(sheet)

    # go through process working time
    attend_records, time_info, projects = processMonths(sheet, salary_records)

    # print("attend_records:", attend_records)
    # print("time_info:", time_info)
    # print("projects:", projects)

    projectCostMap = {}
    departCostMap = {}
    personalMap = {}
    nameNotInTimeInfo = {}
    nameNotInAttendRecords = {}
    prjDepartMap = {}

    for name in salary_records:
        totalAttendHours = sum(attend_records[name])
        departName = salary_records[name]["depart"]

        if departName not in departCostMap:
            departCostMap[departName] = 0.0

        if name not in time_info:
            print("name not in time_info:", name, "put to depart cost")
            nameNotInTimeInfo[name] = 1
            departCost = salary_records[name]["奖金"]
            personalMap[name] = {
                "depart_cost": departCost,
                "project_cost": 0.0,
                "depart_hours": totalAttendHours,
                "project_hours": 0.0,
            }

            departCostMap[departName] += departCost

        else:
            deparHours, projectHours = computePersonalDepartCost(
                time_info[name], attend_records[name])

            projectHours = sum(time_info[name]["total_hours"])

            baseHours = deparHours + projectHours

            perHourCost = salary_records[name]["奖金"] / baseHours

            departCost = deparHours * perHourCost
            personalProjectCost = 0.0
            # go through all projects
            for m in range(0, 12):
                for project in time_info[name]["projects"][m]:
                    if project not in projectCostMap:
                        projectCostMap[project] = 0.0
                    prjCost = time_info[name]["projects"][m][project] * \
                        perHourCost
                    projectCostMap[project] += prjCost
                    personalProjectCost += prjCost
                    if project not in prjDepartMap:
                        prjDepartMap[project] = {}
                    if departName not in prjDepartMap[project]:
                        prjDepartMap[project][departName] = 0.0
                    prjDepartMap[project][departName] += prjCost

            personalMap[name] = {
                "depart_cost": departCost,
                "project_cost": personalProjectCost,
                "depart_hours": deparHours,
                "project_hours": projectHours,
            }

            departCostMap[departName] += departCost

            print("name:", name, "departHours:", deparHours, "totalAttendHours:", totalAttendHours, "projectHours:", sum(time_info[name]["total_hours"]),
                  "perHourCost:", perHourCost, "personalProjectCost", personalProjectCost, "departCost:", departCost, "salary:", salary_records[name]["奖金"])

    print("nameNotInTimeInfo:", nameNotInTimeInfo)

    # write out 2 sheets, one for personal, one for all projects
    personalSheet = sheet.create_sheet("员工汇总")
    titles = ["姓名", "部门", "考勤", "部门小时", "项目小时", "部门成本", "项目成本", "奖金"]
    # print(titles)
    personalSheet.append(titles)
    for name in personalMap:
        item = personalMap[name]
        row = [name, salary_records[name]["depart"], sum(attend_records[name]), item["depart_hours"], item["project_hours"],
               round(item["depart_cost"], 4), round(item["project_cost"], 4), salary_records[name]["奖金"]]
        personalSheet.append(row)

    projectSheet = sheet.create_sheet("项目汇总")
    titles = ["项目编码", "项目名称", "部门", "部门成本"]

    projectSheet.append(titles)
    # for project in projectCostMap:
    #     row = [project, round(projectCostMap[project], 2)]
    #     projectSheet.append(row)
    for project in prjDepartMap:
        for depart in prjDepartMap[project]:
            row = [project, projects_info[project], depart, round(
                prjDepartMap[project][depart], 4)]
            projectSheet.append(row)

    for depart in departCostMap:
        row = [depart, 0, 0, round(departCostMap[depart], 4)]
        projectSheet.append(row)

    sheet.save("year_output.xlsx")


def isEmptyCell(cell):
    return cell.value is None or cell.value == ""


# setup a employee name to depart and salary infos mapping


def setupSalary(wb):
    salary = {}

    salarySheet = wb["2022年年终奖"]
    # print("Salary info length:", salarySheet.max_row)
    # the end row is not used
    for i in range(2, salarySheet.max_row):
        nameCell = salarySheet.cell(row=i, column=1)
        if isEmptyCell(nameCell):
            print("salary empty name cell:", i)
            continue
        # print("process:", nameCell.value)
        departCell = salarySheet.cell(row=i, column=12)
        if isEmptyCell(departCell):
            print("salary empty depart cell:", i)
            continue
        # print("process salary:", nameCell.value, i,
        #      salarySheet.cell(row=1, column=3).value)

        # check if name already exist
        if nameCell.value in salary:
            print("salary name exist:", nameCell.value)
            # panic
            sys.exit(1)

        salary[nameCell.value] = {
            "depart": departCell.value,
            "奖金": float(salarySheet.cell(row=i, column=18).value),
        }

    # print(salary)
    return salary


# compute all employee's attendance hours
# sheet column 11 is attendance days of current month
# return a map of employee name to attendance hours(month)


def processMonths(wb, salary):
    # project id as key
    monthProjects = {}
    missSalaryNames = {}
    monthAttendResult = {}
    monthPersonalWorkResult = {}

    for m in range(1, 13):
        sheetName = str(m) + "月" + "考勤"
        print("process sheet:", sheetName)
        attendSheet = wb[sheetName]
        for i in range(2, attendSheet.max_row + 1):
            nameCell = attendSheet.cell(row=i, column=1)
            name = nameCell.value
            hours = float(attendSheet.cell(row=i, column=11).value) * 8
            # print("attend name:", name, hours)
            if name not in monthAttendResult:
                monthAttendResult[name] = [0.0] * 12
            monthAttendResult[name][m - 1] = hours
        # process this employee's working time record
        sheetName = str(m) + "月" + "工时"
        print("process sheet:", sheetName)
        timeRecordsSheet = wb[sheetName]
        for i in range(2, timeRecordsSheet.max_row + 1):
            nameCell = timeRecordsSheet.cell(row=i, column=2)
            if isEmptyCell(nameCell):
                print("empty name cell:", i)
                continue
            name = nameCell.value
            if name not in salary:
                print("name not in salary", name)
                missSalaryNames[name] = 1
                continue
            projectIdCell = timeRecordsSheet.cell(row=i, column=8)
            if isEmptyCell(projectIdCell):
                print("empty project cell:", i, name)
                continue

            if name not in monthPersonalWorkResult:
                monthPersonalWorkResult[name] = {
                    # init 12 elements to zero
                    "total_hours": [0.0] * 12,
                    "projects": [{}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}],
                }
            hoursCell = timeRecordsSheet.cell(row=i, column=12)
            if isEmptyCell(hoursCell):
                print("empty hours cell:", i)
                continue
            hours = int(hoursCell.value)
            monthPersonalWorkResult[name]["total_hours"][m - 1] += hours
            projectNameCell = timeRecordsSheet.cell(row=i, column=9)
            if projectIdCell.value not in projects_info:
                projects_info[projectIdCell.value] = projectNameCell.value

            if projectIdCell.value not in monthPersonalWorkResult[name]["projects"][m - 1]:
                monthPersonalWorkResult[name]["projects"][m -
                                                          1][projectIdCell.value] = 0.0

            monthPersonalWorkResult[name]["projects"][m -
                                                      1][projectIdCell.value] += hours

            if projectIdCell.value not in monthProjects:
                # employee name as key (should use employee id, TODO)
                monthProjects[projectIdCell.value] = [
                    {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}, {}]

            if name not in monthProjects[projectIdCell.value][m - 1]:
                monthProjects[projectIdCell.value][m - 1][name] = 0
            monthProjects[projectIdCell.value][m - 1][name] += hours

    print("miss salary names:", missSalaryNames.keys())
    return (monthAttendResult, monthPersonalWorkResult, monthProjects)


def computePersonalDepartCost(timeRecords, attendRecords):
    depart_hours = 0.0
    project_hours = 0.0
    # go thought 12 months
    for m in range(0, 12):
        monthProjectHours = timeRecords["total_hours"][m]
        monthAttendHours = attendRecords[m]
        if monthAttendHours > monthProjectHours:
            # treat diff as depart cost

            diff = monthAttendHours - monthProjectHours

            print("monthAttendHours:", monthAttendHours,
                  "monthProjectHours:", monthProjectHours, "diff:", diff)

            depart_hours += diff
        project_hours += monthProjectHours
    return (depart_hours, project_hours)


init()
